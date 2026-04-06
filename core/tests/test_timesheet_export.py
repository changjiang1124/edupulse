from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from core.models import TeacherAttendance
from facilities.models import Facility


@override_settings(SECURE_SSL_REDIRECT=False)
class TimesheetExportViewTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.admin = user_model.objects.create_user(
            username='legacy-admin',
            password='Admin123!',
            first_name='Legacy',
            last_name='Admin',
            role='admin',
            is_active=True,
            is_active_staff=True,
        )
        self.teacher = user_model.objects.create_user(
            username='legacy-teacher',
            password='Teacher123!',
            first_name='Legacy',
            last_name='Teacher',
            role='teacher',
            is_active=True,
            is_active_staff=True,
        )
        self.office_admin = user_model.objects.create_user(
            username='legacy-office',
            password='Office123!',
            first_name='Office',
            last_name='Manager',
            role='admin',
            is_active=True,
            is_active_staff=True,
        )
        self.facility = Facility.objects.create(
            name='Legacy Export Studio',
            address='1 Legacy Way',
            latitude=Decimal('0.0'),
            longitude=Decimal('0.0'),
            attendance_radius=500,
            is_active=True,
        )
        self.client = Client()

    def _create_manual_session(self, staff_member, start, end, reason):
        for clock_type, timestamp in (('clock_in', start), ('clock_out', end)):
            TeacherAttendance.objects.create(
                teacher=staff_member,
                clock_type=clock_type,
                source='manual',
                timestamp=timestamp,
                facility=self.facility,
                manual_reason=reason,
                notes='Imported from paper record.',
                created_by=self.admin,
                updated_by=self.admin,
            )

    def test_timesheet_page_export_link_keeps_current_date_range(self):
        self.client.login(username='legacy-teacher', password='Teacher123!')

        response = self.client.get(
            reverse('timesheet'),
            {'start_date': '2026-03-01', 'end_date': '2026-03-31'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f"{reverse('timesheet_export')}?start_date=2026-03-01&end_date=2026-03-31",
        )

    def test_legacy_export_page_prefills_current_filters(self):
        self.client.login(username='legacy-admin', password='Admin123!')

        response = self.client.get(
            reverse('timesheet_export'),
            {'start_date': '2026-03-01', 'end_date': '2026-03-31', 'staff_id': str(self.office_admin.pk)},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'value="2026-03-01"')
        self.assertContains(response, 'value="2026-03-31"')
        self.assertContains(response, 'All Active Staff')
        self.assertContains(response, f'value="{self.office_admin.pk}" selected')

    def test_admin_can_export_non_teacher_staff_from_legacy_route(self):
        self.client.login(username='legacy-admin', password='Admin123!')

        start = timezone.localtime(timezone.now()).replace(second=0, microsecond=0) - timedelta(hours=4)
        end = start + timedelta(hours=2)
        self._create_manual_session(
            self.office_admin,
            start,
            end,
            'Payroll correction for office shift.',
        )

        response = self.client.post(
            reverse('timesheet_export'),
            {
                'staff_id': str(self.office_admin.pk),
                'start_date': timezone.localtime(start).strftime('%Y-%m-%d'),
                'end_date': timezone.localtime(end).strftime('%Y-%m-%d'),
                'format': 'excel',
            },
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content))
        worksheet = workbook.active
        flattened_values = [
            str(cell)
            for row in worksheet.iter_rows(values_only=True)
            for cell in row
            if cell not in (None, '')
        ]

        self.assertIn('Office Manager', flattened_values)
        self.assertIn('Payroll correction for office shift.', flattened_values)
        self.assertIn('Manual Entry', flattened_values)

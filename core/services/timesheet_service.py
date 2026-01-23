"""
Timesheet Export Service for EduPulse
Handles generation of timesheet reports in Excel format
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Q
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Any, Optional
import logging

logger = logging.getLogger(__name__)


class TimesheetExportService:
    """
    Service for generating Excel timesheet reports
    """
    
    @staticmethod
    def export_teacher_timesheet(teacher=None, start_date=None, end_date=None, format='excel'):
        """
        Export teacher timesheet data to Excel format
        
        Args:
            teacher: Specific teacher (if None, exports all teachers)
            start_date: Start date for report
            end_date: End date for report  
            format: Export format ('excel' or 'csv')
            
        Returns:
            HttpResponse with Excel file
        """
        try:
            from accounts.models import Staff
            from core.services.staff_timesheet_service import StaffTimesheetService
            
            # Set default date range if not provided
            if not end_date:
                end_date = timezone.now().date()
            if not start_date:
                start_date = end_date - timedelta(days=30)  # Last 30 days
            
            # Get staff list for export
            if teacher:
                staff_queryset = [teacher]
            else:
                staff_queryset = list(
                    Staff.objects.filter(role='teacher', is_active=True)
                    .order_by('first_name', 'last_name')
                )

            staff_timesheets = []
            for staff_member in staff_queryset:
                timesheet_data = StaffTimesheetService.get_staff_timesheet_data(
                    staff_member, start_date, end_date
                )
                paired_records = timesheet_data.get('paired_records', [])
                if paired_records or teacher:
                    staff_timesheets.append({
                        'staff': staff_member,
                        'paired_records': paired_records
                    })
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Set worksheet title
            teacher_name = f" - {teacher.get_full_name()}" if teacher else ""
            ws.title = f"Timesheet{teacher_name}"[:31]  # Excel sheet name limit
            
            # Generate the timesheet
            TimesheetExportService._generate_timesheet_worksheet(
                ws, staff_timesheets, start_date, end_date, teacher
            )
            
            # Prepare response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Set filename
            date_str = f"{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}"
            filename = f"timesheet_{teacher_name.replace(' - ', '_').replace(' ', '_').lower()}_{date_str}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Save workbook to response
            wb.save(response)
            
            logger.info(f"Timesheet exported successfully for period {start_date} to {end_date}")
            return response
            
        except Exception as e:
            logger.error(f"Error exporting timesheet: {str(e)}")
            raise e
    
    @staticmethod
    def _generate_timesheet_worksheet(ws, staff_timesheets, start_date, end_date, teacher):
        """
        Generate the timesheet worksheet with data and formatting
        """
        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Data styling
        data_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title and header
        ws['A1'] = 'Perth Art School - Teacher Timesheet'
        ws.merge_cells('A1:H1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Report details
        ws['A3'] = f'Report Period: {start_date.strftime("%d/%m/%Y")} - {end_date.strftime("%d/%m/%Y")}'
        if teacher:
            ws['A4'] = f'Teacher: {teacher.get_full_name()}'
        ws['A5'] = f'Generated: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        
        # Column headers
        headers = [
            'Date', 'Teacher', 'Clock In', 'Clock Out', 'Facility',
            'Classes', 'Duration (Hours)', 'Notes'
        ]
        
        header_row = 7
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_alignment
        
        # Data rows
        row = header_row + 1
        daily_hours = {}  # Track daily hours by teacher and date

        for staff_entry in staff_timesheets:
            staff_member = staff_entry['staff']
            paired_records = staff_entry.get('paired_records', [])
            daily_hours[staff_member.id] = {}

            for record in paired_records:
                duration_hours = record.get('duration_hours')

                date_key = record['date']
                if date_key not in daily_hours[staff_member.id]:
                    daily_hours[staff_member.id][date_key] = Decimal('0')
                if duration_hours is not None:
                    daily_hours[staff_member.id][date_key] += Decimal(str(duration_hours))

                class_names = ', '.join([
                    cls.course.name for cls in record.get('classes', [])
                ]) if record.get('classes') else 'General'

                clock_in_time = (
                    record.get('clock_in_time').strftime('%H:%M')
                    if record.get('clock_in_time') else '--'
                )
                clock_out_time = (
                    record.get('clock_out_time').strftime('%H:%M')
                    if record.get('clock_out_time') else '--'
                )
                duration_value = ''
                if duration_hours is not None:
                    duration_value = float(
                        Decimal(str(duration_hours)).quantize(
                            Decimal('0.01'), rounding=ROUND_HALF_UP
                        )
                    )

                data_row = [
                    record['date'].strftime('%d/%m/%Y'),
                    staff_member.get_full_name(),
                    clock_in_time,
                    clock_out_time,
                    record.get('facility').name if record.get('facility') else 'N/A',
                    class_names,
                    duration_value,
                    record.get('notes') or ''
                ]

                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = data_border

                    if col == 7 and value != '':  # Duration column
                        cell.alignment = Alignment(horizontal='right')

                row += 1

            row = TimesheetExportService._add_teacher_summary(
                ws, row, staff_member, daily_hours
            )
            row += 1  # Space between teachers

        # Add overall summary
        row += 1
        TimesheetExportService._add_overall_summary(ws, row, staff_timesheets, daily_hours)
        
        # Adjust column widths
        column_widths = [12, 20, 15, 10, 20, 25, 15, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    @staticmethod
    def _calculate_attendance_duration(attendance):
        """
        Calculate duration for attendance record based on clock in/out pairs
        """
        try:
            from core.models import TeacherAttendance
            
            # For clock_out, find matching clock_in
            if attendance.clock_type == 'clock_out':
                # Find the most recent clock_in for the same teacher on the same day
                clock_in = TeacherAttendance.objects.filter(
                    teacher=attendance.teacher,
                    clock_type='clock_in',
                    timestamp__date=attendance.timestamp.date(),
                    timestamp__lt=attendance.timestamp
                ).order_by('-timestamp').first()
                
                if clock_in:
                    duration = attendance.timestamp - clock_in.timestamp
                    hours = Decimal(str(duration.total_seconds() / 3600)).quantize(
                        Decimal('0.01'), rounding=ROUND_HALF_UP
                    )
                    return hours
            
            return None
            
        except Exception as e:
            logger.warning(f"Error calculating duration for attendance {attendance.id}: {str(e)}")
            return None
    
    @staticmethod
    def _add_teacher_summary(ws, row, teacher, daily_hours):
        """
        Add summary row for a teacher
        """
        teacher_total = sum(
            sum(daily_hours.get(teacher.id, {}).values(), Decimal('0')),
        )
        
        # Teacher summary row
        ws.merge_cells(f'A{row}:E{row}')
        summary_cell = ws[f'A{row}']
        summary_cell.value = f'Total for {teacher.get_full_name()}:'
        summary_cell.font = Font(bold=True)
        summary_cell.alignment = Alignment(horizontal='right')
        
        ws[f'G{row}'] = float(teacher_total)
        ws[f'G{row}'].font = Font(bold=True)
        ws[f'G{row}'].alignment = Alignment(horizontal='right')
        
        return row + 1
    
    @staticmethod
    def _add_overall_summary(ws, row, staff_timesheets, daily_hours):
        """
        Add overall summary section
        """
        # Summary header
        ws[f'A{row}'] = 'SUMMARY'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2
        
        # Calculate totals
        total_teachers = len(staff_timesheets)
        total_hours = sum(
            sum(teacher_daily.values(), Decimal('0'))
            for teacher_daily in daily_hours.values()
        )
        total_sessions = sum(
            len(entry.get('paired_records', [])) for entry in staff_timesheets
        )
        incomplete_sessions = sum(
            1
            for entry in staff_timesheets
            for record in entry.get('paired_records', [])
            if not record.get('is_complete')
        )
        
        # Summary data
        summary_data = [
            ('Total Teachers:', total_teachers),
            ('Total Hours:', float(total_hours)),
            ('Total Sessions:', total_sessions),
            ('Incomplete Sessions:', incomplete_sessions),
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
    
    @staticmethod
    def generate_monthly_summary(year, month):
        """
        Generate monthly summary report for all teachers
        """
        try:
            from accounts.models import Staff
            from core.services.staff_timesheet_service import StaffTimesheetService
            from datetime import date
            import calendar
            
            # Get first and last day of month
            start_date = date(year, month, 1)
            last_day = calendar.monthrange(year, month)[1]
            end_date = date(year, month, last_day)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Monthly Summary {year}-{month:02d}"
            
            # Process data
            teacher_data = {}
            staff_queryset = Staff.objects.filter(
                role='teacher',
                is_active=True
            ).order_by('first_name', 'last_name')

            for staff_member in staff_queryset:
                timesheet_data = StaffTimesheetService.get_staff_timesheet_data(
                    staff_member, start_date, end_date
                )
                paired_records = timesheet_data.get('paired_records', [])
                if not paired_records:
                    continue

                total_hours = sum(
                    Decimal(str(record['duration_hours']))
                    for record in paired_records
                    if record.get('duration_hours') is not None
                )
                days_worked = {
                    record['date']
                    for record in paired_records
                    if record.get('duration_hours') is not None
                }

                teacher_data[staff_member.id] = {
                    'teacher': staff_member,
                    'total_hours': total_hours,
                    'days_worked': days_worked,
                    'total_sessions': len(paired_records)
                }
            
            # Generate worksheet
            TimesheetExportService._generate_monthly_summary_worksheet(
                ws, teacher_data, start_date, end_date
            )
            
            # Prepare response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            filename = f"monthly_summary_{year}_{month:02d}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f"Error generating monthly summary: {str(e)}")
            raise e
    
    @staticmethod
    def _generate_monthly_summary_worksheet(ws, teacher_data, start_date, end_date):
        """
        Generate monthly summary worksheet
        """
        # Header
        ws['A1'] = f'Perth Art School - Monthly Summary'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f'Period: {start_date.strftime("%B %Y")}'
        ws['A3'] = f'Generated: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        
        # Column headers
        headers = ['Teacher', 'Total Hours', 'Days Worked', 'Avg Hours/Day', 'Total Sessions']
        header_row = 5
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data rows
        row = header_row + 1
        total_hours_all = Decimal('0')
        
        for teacher_info in teacher_data.values():
            teacher = teacher_info['teacher']
            total_hours = teacher_info['total_hours']
            days_worked = len(teacher_info['days_worked'])
            avg_hours = float(total_hours / days_worked) if days_worked > 0 else 0
            total_sessions = teacher_info['total_sessions']
            
            # Add row data
            ws.cell(row=row, column=1, value=teacher.get_full_name())
            ws.cell(row=row, column=2, value=float(total_hours))
            ws.cell(row=row, column=3, value=days_worked)
            ws.cell(row=row, column=4, value=round(avg_hours, 2))
            ws.cell(row=row, column=5, value=total_sessions)
            
            total_hours_all += total_hours
            row += 1
        
        # Totals row
        row += 1
        ws.cell(row=row, column=1, value='TOTAL:').font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(total_hours_all)).font = Font(bold=True)
        
        # Adjust column widths
        column_widths = [25, 15, 15, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

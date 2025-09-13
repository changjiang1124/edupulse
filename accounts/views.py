from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView, View
from django.urls import reverse_lazy
from django.db.models import Q

from .models import Staff
from .forms import StaffForm, StaffCreationForm


class AdminRequiredMixin(UserPassesTestMixin):
    """Admin permission check mixin"""
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.role == 'admin'


class ProfileView(LoginRequiredMixin, DetailView):
    """User profile view - shows current user's staff profile"""
    model = Staff
    template_name = 'core/staff/detail.html'
    context_object_name = 'staff'
    
    def get_object(self):
        # Always return the current user
        return self.request.user
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add teacher-related courses and classes - with try/except to avoid errors
        try:
            # Import here to avoid circular imports
            from academics.models import Course, Class
            from core.services.staff_timesheet_service import StaffTimesheetService
            from datetime import datetime, timedelta
            
            context['taught_courses'] = Course.objects.filter(teacher=self.object, status='published')
            context['taught_classes'] = Class.objects.filter(
                course__teacher=self.object
            ).select_related('course').order_by('-date')[:5]
            
            # Add timesheet data
            start_date = self.request.GET.get('timesheet_start')
            end_date = self.request.GET.get('timesheet_end')
            
            # Parse dates if provided
            if start_date:
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                except ValueError:
                    start_date = None
            if end_date:
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                except ValueError:
                    end_date = None
            
            # Get timesheet data
            timesheet_data = StaffTimesheetService.get_staff_timesheet_data(
                self.object, start_date, end_date
            )
            
            context['timesheet_data'] = timesheet_data
            context['timesheet_start_date'] = timesheet_data['date_range']['start_date']
            context['timesheet_end_date'] = timesheet_data['date_range']['end_date']
            
        except ImportError:
            context['taught_courses'] = []
            context['taught_classes'] = []
            context['timesheet_data'] = {'paired_records': [], 'summary': {}}
        # Add context flag to indicate this is profile view
        context['is_profile_view'] = True
        return context


class StaffListView(AdminRequiredMixin, ListView):
    model = Staff
    template_name = 'core/staff/list.html'
    context_object_name = 'staff_list'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Staff.objects.filter(is_active_staff=True)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(username__icontains=search)
            )
        return queryset.order_by('last_name', 'first_name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        return context


class StaffCreateView(AdminRequiredMixin, CreateView):
    model = Staff
    form_class = StaffCreationForm
    template_name = 'core/staff/form.html'
    success_url = reverse_lazy('accounts:staff_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'Staff member {form.instance.first_name} {form.instance.last_name} created successfully!')
        return super().form_valid(form)


class StaffDetailView(AdminRequiredMixin, DetailView):
    model = Staff
    template_name = 'core/staff/detail.html'
    context_object_name = 'staff'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add teacher-related courses and classes - with try/except to avoid errors
        try:
            # Import here to avoid circular imports
            from academics.models import Course, Class
            from core.services.staff_timesheet_service import StaffTimesheetService
            from datetime import datetime, timedelta
            
            context['taught_courses'] = Course.objects.filter(teacher=self.object, status='published')
            context['taught_classes'] = Class.objects.filter(
                course__teacher=self.object
            ).select_related('course').order_by('-date')[:5]
            
            # Add timesheet data
            start_date = self.request.GET.get('timesheet_start')
            end_date = self.request.GET.get('timesheet_end')
            
            # Parse dates if provided
            if start_date:
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                except ValueError:
                    start_date = None
            if end_date:
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                except ValueError:
                    end_date = None
            
            # Get timesheet data
            timesheet_data = StaffTimesheetService.get_staff_timesheet_data(
                self.object, start_date, end_date
            )
            
            context['timesheet_data'] = timesheet_data
            context['timesheet_start_date'] = timesheet_data['date_range']['start_date']
            context['timesheet_end_date'] = timesheet_data['date_range']['end_date']
            
        except ImportError:
            context['taught_courses'] = []
            context['taught_classes'] = []
            context['timesheet_data'] = {'paired_records': [], 'summary': {}}
        return context


class StaffUpdateView(AdminRequiredMixin, UpdateView):
    model = Staff
    form_class = StaffForm
    template_name = 'core/staff/form.html'
    
    def get_success_url(self):
        return reverse_lazy('accounts:staff_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, f'Staff member {form.instance.first_name} {form.instance.last_name} updated successfully!')
        return super().form_valid(form)


class StaffTimesheetExportView(LoginRequiredMixin, View):
    """Export staff timesheet data in CSV or Excel format"""
    
    def get(self, request, pk):
        import csv
        from django.http import HttpResponse
        from datetime import datetime
        from io import BytesIO
        
        # Get staff member
        staff = get_object_or_404(Staff, pk=pk)
        
        # Permission check - only admin or the staff member themselves
        if not (request.user.role == 'admin' or request.user == staff):
            messages.error(request, 'You do not have permission to view this timesheet.')
            return redirect('core:dashboard')
        
        # Get parameters
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        export_format = request.GET.get('format', 'csv').lower()
        
        # Parse dates
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                start_date = None
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                end_date = None
        
        # Get timesheet data
        try:
            from core.services.staff_timesheet_service import StaffTimesheetService
            timesheet_data = StaffTimesheetService.get_staff_timesheet_data(
                staff, start_date, end_date
            )
        except ImportError:
            messages.error(request, 'Timesheet service is not available.')
            return redirect('accounts:staff_detail', pk=pk)
        
        # Generate filename
        date_range = f"{timesheet_data['date_range']['start_date'].strftime('%Y%m%d')}-{timesheet_data['date_range']['end_date'].strftime('%Y%m%d')}"
        filename = f"{staff.first_name}_{staff.last_name}_timesheet_{date_range}"
        
        if export_format == 'excel':
            return self._generate_excel(timesheet_data, staff, filename)
        else:
            return self._generate_csv(timesheet_data, staff, filename)
    
    def _generate_csv(self, timesheet_data, staff, filename):
        """Generate CSV export"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.writer(response)
        
        # Header information
        writer.writerow(['Staff Timesheet Export'])
        writer.writerow(['Name:', f"{staff.first_name} {staff.last_name}"])
        writer.writerow(['Email:', staff.email or 'N/A'])
        writer.writerow(['Date Range:', f"{timesheet_data['date_range']['start_date']} to {timesheet_data['date_range']['end_date']}"])
        writer.writerow(['Export Date:', datetime.now().strftime('%Y-%m-%d %H:%M')])
        writer.writerow([])  # Empty row
        
        # Summary
        if timesheet_data['summary']:
            summary = timesheet_data['summary']
            writer.writerow(['SUMMARY'])
            writer.writerow(['Total Hours:', f"{summary.get('total_hours', 0)}h"])
            writer.writerow(['Total Days:', summary.get('total_days', 0)])
            writer.writerow(['Completed Sessions:', summary.get('completed_sessions', 0)])
            writer.writerow(['Average Hours per Day:', f"{summary.get('avg_hours_per_day', 0)}h"])
            writer.writerow([])  # Empty row
        
        # Detailed records
        writer.writerow(['DETAILED TIMESHEET'])
        writer.writerow(['Date', 'Clock In', 'Clock Out', 'Classes', 'Facility', 'Hours'])
        
        for record in timesheet_data.get('paired_records', []):
            classes_text = ', '.join([cls.course.name for cls in record.get('classes', [])])
            if not classes_text:
                classes_text = 'General Work'
            
            facility_name = record.get('facility').name if record.get('facility') else 'N/A'
            duration = record.get('duration_hours', 0)
            
            writer.writerow([
                record['date'].strftime('%Y-%m-%d'),
                record.get('clock_in_time').strftime('%H:%M') if record.get('clock_in_time') else '--',
                record.get('clock_out_time').strftime('%H:%M') if record.get('clock_out_time') else '--',
                classes_text,
                facility_name,
                f"{duration}h" if duration else '--'
            ])
        
        return response
    
    def _generate_excel(self, timesheet_data, staff, filename):
        """Generate Excel export"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from django.http import HttpResponse
            from io import BytesIO
        except ImportError:
            messages.error(self.request, 'Excel export requires openpyxl. Please install it.')
            return redirect('accounts:staff_detail', pk=staff.pk)
        
        # Create workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Timesheet'
        
        # Styles
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=10)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
        
        row = 1
        
        # Header information
        worksheet[f'A{row}'] = 'Staff Timesheet Export'
        worksheet[f'A{row}'].font = header_font
        row += 2
        
        worksheet[f'A{row}'] = 'Name:'
        worksheet[f'B{row}'] = f"{staff.first_name} {staff.last_name}"
        worksheet[f'A{row}'].font = subheader_font
        row += 1
        
        worksheet[f'A{row}'] = 'Email:'
        worksheet[f'B{row}'] = staff.email or 'N/A'
        worksheet[f'A{row}'].font = subheader_font
        row += 1
        
        worksheet[f'A{row}'] = 'Date Range:'
        worksheet[f'B{row}'] = f"{timesheet_data['date_range']['start_date']} to {timesheet_data['date_range']['end_date']}"
        worksheet[f'A{row}'].font = subheader_font
        row += 1
        
        worksheet[f'A{row}'] = 'Export Date:'
        worksheet[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        worksheet[f'A{row}'].font = subheader_font
        row += 2
        
        # Summary
        if timesheet_data['summary']:
            summary = timesheet_data['summary']
            worksheet[f'A{row}'] = 'SUMMARY'
            worksheet[f'A{row}'].font = header_font
            row += 1
            
            worksheet[f'A{row}'] = 'Total Hours:'
            worksheet[f'B{row}'] = f"{summary.get('total_hours', 0)}h"
            row += 1
            
            worksheet[f'A{row}'] = 'Total Days:'
            worksheet[f'B{row}'] = summary.get('total_days', 0)
            row += 1
            
            worksheet[f'A{row}'] = 'Completed Sessions:'
            worksheet[f'B{row}'] = summary.get('completed_sessions', 0)
            row += 1
            
            worksheet[f'A{row}'] = 'Average Hours per Day:'
            worksheet[f'B{row}'] = f"{summary.get('avg_hours_per_day', 0)}h"
            row += 2
        
        # Detailed records header
        worksheet[f'A{row}'] = 'DETAILED TIMESHEET'
        worksheet[f'A{row}'].font = header_font
        row += 1
        
        # Table headers
        headers = ['Date', 'Clock In', 'Clock Out', 'Classes', 'Facility', 'Hours']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col, value=header)
            cell.font = subheader_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Data rows
        for record in timesheet_data.get('paired_records', []):
            classes_text = ', '.join([cls.course.name for cls in record.get('classes', [])])
            if not classes_text:
                classes_text = 'General Work'
            
            facility_name = record.get('facility').name if record.get('facility') else 'N/A'
            duration = record.get('duration_hours', 0)
            
            data = [
                record['date'].strftime('%Y-%m-%d'),
                record.get('clock_in_time').strftime('%H:%M') if record.get('clock_in_time') else '--',
                record.get('clock_out_time').strftime('%H:%M') if record.get('clock_out_time') else '--',
                classes_text,
                facility_name,
                f"{duration}h" if duration else '--'
            ]
            
            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [1, 2, 3, 6]:  # Date, times, and hours columns
                    cell.alignment = Alignment(horizontal='center')
            row += 1
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        return response


class StaffTimesheetOverviewView(AdminRequiredMixin, ListView):
    """Timesheet overview for all staff members"""
    template_name = 'core/staff/timesheet_overview.html'
    context_object_name = 'staff_timesheets'
    paginate_by = 20
    
    def get_queryset(self):
        # Get all active staff members
        return Staff.objects.filter(is_active_staff=True, role='teacher').order_by('last_name', 'first_name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get date range from request
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        # Parse dates if provided
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                start_date = None
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                end_date = None
        
        # Get timesheet data for all staff
        try:
            from core.services.staff_timesheet_service import StaffTimesheetService
            from datetime import datetime, timedelta
            
            overview_data = StaffTimesheetService.get_all_staff_timesheet_data(
                self.get_queryset(), start_date, end_date
            )
            
            context['timesheet_overview'] = overview_data
            context['start_date'] = overview_data['date_range']['start_date']
            context['end_date'] = overview_data['date_range']['end_date']
            
        except ImportError:
            context['timesheet_overview'] = {
                'overall_summary': {},
                'staff_summaries': [],
                'date_range': {'start_date': None, 'end_date': None}
            }
        
        return context


class StaffTimesheetOverviewExportView(AdminRequiredMixin, View):
    """Export all staff timesheet data in CSV or Excel format"""
    
    def get(self, request):
        import csv
        from django.http import HttpResponse
        from datetime import datetime
        from io import BytesIO
        
        # Get parameters
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        export_format = request.GET.get('format', 'csv').lower()
        
        # Parse dates
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                start_date = None
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                end_date = None
        
        # Get all active staff
        staff_queryset = Staff.objects.filter(is_active_staff=True, role='teacher').order_by('last_name', 'first_name')
        
        # Get timesheet overview data
        try:
            from core.services.staff_timesheet_service import StaffTimesheetService
            overview_data = StaffTimesheetService.get_all_staff_timesheet_data(
                staff_queryset, start_date, end_date
            )
        except ImportError:
            messages.error(request, 'Timesheet service is not available.')
            return redirect('accounts:staff_timesheet_overview')
        
        # Generate filename
        date_range = f"{overview_data['date_range']['start_date'].strftime('%Y%m%d')}-{overview_data['date_range']['end_date'].strftime('%Y%m%d')}"
        filename = f"all_staff_timesheet_{date_range}"
        
        if export_format == 'excel':
            return self._generate_excel(overview_data, filename)
        else:
            return self._generate_csv(overview_data, filename)
    
    def _generate_csv(self, overview_data, filename):
        """Generate CSV export for all staff"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.writer(response)
        
        # Header information
        writer.writerow(['All Staff Timesheet Export'])
        writer.writerow(['Date Range:', f"{overview_data['date_range']['start_date']} to {overview_data['date_range']['end_date']}"])
        writer.writerow(['Export Date:', datetime.now().strftime('%Y-%m-%d %H:%M')])
        writer.writerow([])  # Empty row
        
        # Overall summary
        if overview_data['overall_summary']:
            summary = overview_data['overall_summary']
            writer.writerow(['OVERALL SUMMARY'])
            writer.writerow(['Total Hours:', f"{summary.get('total_hours', 0)}h"])
            writer.writerow(['Active Staff Count:', summary.get('active_staff_count', 0)])
            writer.writerow(['Total Sessions:', summary.get('total_sessions', 0)])
            writer.writerow(['Average Hours per Staff:', f"{summary.get('average_hours_per_staff', 0)}h"])
            writer.writerow([])  # Empty row
        
        # Staff summary
        writer.writerow(['STAFF SUMMARY'])
        writer.writerow(['Staff Name', 'Total Hours', 'Working Days', 'Sessions', 'Average Hours/Day'])
        
        for staff_data in overview_data.get('staff_summaries', []):
            writer.writerow([
                f"{staff_data['staff'].first_name} {staff_data['staff'].last_name}",
                f"{staff_data.get('total_hours', 0)}h",
                staff_data.get('working_days', 0),
                staff_data.get('sessions', 0),
                f"{staff_data.get('average_hours_per_day', 0)}h"
            ])
        
        return response
    
    def _generate_excel(self, overview_data, filename):
        """Generate Excel export for all staff"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from django.http import HttpResponse
            from io import BytesIO
        except ImportError:
            messages.error(self.request, 'Excel export requires openpyxl. Please install it.')
            return redirect('accounts:staff_timesheet_overview')
        
        # Create workbook
        workbook = openpyxl.Workbook()
        
        # Summary sheet
        summary_sheet = workbook.active
        summary_sheet.title = 'Summary'
        
        # Styles
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=10)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
        
        row = 1
        
        # Header information
        summary_sheet[f'A{row}'] = 'All Staff Timesheet Export'
        summary_sheet[f'A{row}'].font = header_font
        row += 2
        
        summary_sheet[f'A{row}'] = 'Date Range:'
        summary_sheet[f'B{row}'] = f"{overview_data['date_range']['start_date']} to {overview_data['date_range']['end_date']}"
        summary_sheet[f'A{row}'].font = subheader_font
        row += 1
        
        summary_sheet[f'A{row}'] = 'Export Date:'
        summary_sheet[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        summary_sheet[f'A{row}'].font = subheader_font
        row += 2
        
        # Overall summary
        if overview_data['overall_summary']:
            summary = overview_data['overall_summary']
            summary_sheet[f'A{row}'] = 'OVERALL SUMMARY'
            summary_sheet[f'A{row}'].font = header_font
            row += 1
            
            summary_sheet[f'A{row}'] = 'Total Hours:'
            summary_sheet[f'B{row}'] = f"{summary.get('total_hours', 0)}h"
            row += 1
            
            summary_sheet[f'A{row}'] = 'Active Staff Count:'
            summary_sheet[f'B{row}'] = summary.get('active_staff_count', 0)
            row += 1
            
            summary_sheet[f'A{row}'] = 'Total Sessions:'
            summary_sheet[f'B{row}'] = summary.get('total_sessions', 0)
            row += 1
            
            summary_sheet[f'A{row}'] = 'Average Hours per Staff:'
            summary_sheet[f'B{row}'] = f"{summary.get('average_hours_per_staff', 0)}h"
            row += 2
        
        # Staff summary table
        summary_sheet[f'A{row}'] = 'STAFF SUMMARY'
        summary_sheet[f'A{row}'].font = header_font
        row += 1
        
        # Table headers
        headers = ['Staff Name', 'Total Hours', 'Working Days', 'Sessions', 'Average Hours/Day']
        for col, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=row, column=col, value=header)
            cell.font = subheader_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Staff data
        for staff_data in overview_data.get('staff_summaries', []):
            data = [
                f"{staff_data['staff'].first_name} {staff_data['staff'].last_name}",
                f"{staff_data.get('total_hours', 0)}h",
                staff_data.get('working_days', 0),
                staff_data.get('sessions', 0),
                f"{staff_data.get('average_hours_per_day', 0)}h"
            ]
            
            for col, value in enumerate(data, 1):
                cell = summary_sheet.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [2, 3, 4, 5]:  # Numeric columns
                    cell.alignment = Alignment(horizontal='center')
            row += 1
        
        # Auto-adjust column widths
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        return response
